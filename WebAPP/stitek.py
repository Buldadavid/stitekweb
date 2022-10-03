from PIL import Image, ImageFont, ImageDraw
from openpyxl import load_workbook
import qrcode
import csv
import datetime
from datetime import date

def stitek():
    data = []

    with open("data/data.csv") as file:
        reader = csv.DictReader(file)

        for r in reader:
            data.append({"QR": r["QR"],"MLFB": r["MLFB"], "cislo": r["cislo"],"z": r["z"], "brzda": r["brzda"], "odmer": r["odmer"], "vaha": r["vaha"], "dat_v": r["dat_v"], "tep": r["tep"], "barvaO": r["barvaO"], "barvaZ": r["barvaZ"]})

    #print((data[0])["MLFB"])
    QR = (data[0])["QR"]
    MLFB = (data[0])["MLFB"]
    cislo = (data[0])["cislo"]
    Z = (data[0])["z"]
    dat_v = (data[0])["dat_v"]
    odmer = (data[0])["odmer"]
    brzda = (data[0])["brzda"]
    vaha = (data[0])["vaha"]
    tep = (data[0])["tep"]
    barvaO = (data[0])["barvaO"]
    barvaZ = (data[0])["barvaZ"]

    if Z == "/" :
        Z = ""
        
    else:
        Z = f"Z:{Z}"
        
    
    current_time = datetime.datetime.now()
    dat = str(current_time.year) + "." + str(current_time.month) + "." + str(current_time.day)
    vibroOK = dat + "_" + MLFB + "_" + cislo + "_16,67Hz_OK_VIB12"
    vibroNOK = dat + "_" + MLFB + "_" + cislo + "_16,67Hz_NOK_VIB12"

    todays_date = date.today()
    year = todays_date.year
    q= year-int(dat_v[-4:])
    barva = "#fefefe"

    if q > 2 :
        barva = "#ff0000"
    elif q == 1:
        barva = "yellow"
    elif q == 0:
        barva = "#74ec3a"

    qr_text = "1P"+MLFB+"+SYF"+cislo.replace("-","")

    input_data = qr_text
    qr = qrcode.QRCode(
        version=1,
        box_size=2,
        border=1)
    qr.add_data(input_data)
    qr.make(fit=True)
    im2 = qr.make_image(fill='black', back_color='white')
    my_image = Image.open("tisk-stitku/st.jpg")
    title_font = ImageFont.truetype("tisk-stitku/Roboto-Medium.ttf", 20)

    image_editable = ImageDraw.Draw(my_image)
    image_editable.text((25,35), "3 ~ Motor " + MLFB, (0, 0, 0), font=title_font)
    image_editable.text((25,73), "No. YF " + cislo, (0, 0, 0), font=title_font)
    image_editable.text((340,73), Z, (0, 0, 0), font=title_font)
    image_editable.text((25,105), "Enc. " + odmer, (0, 0, 0), font=title_font)
    image_editable.text((25,135), "Brake "+ brzda, (0, 0, 0), font=title_font)
    image_editable.text((25,175), "m "+ vaha +"Kg", (0, 0, 0), font=title_font)
    image_editable.text((225,105), "Tepl. čidlo " + tep, (0, 0, 0), font=title_font)

    my_image.paste(im2, (340, 135))
    my_image.save("static/pillow_paste.jpg", quality=95)
    print("OK")
    return(QR,MLFB,Z,cislo,odmer,dat_v,barva,vibroOK,vibroNOK,barvaO,barvaZ)

def vyroba(QR):
    wb = load_workbook(filename = 'data/data.xlsx')
    ws = wb['nic']
    
    imput = QR[2:50]
    cd = len(imput)
    #print(cd)
    
#    if " " in QR :
#        return render_template('error/error_qr.html',QR=QR)        

    if cd >= 37 :
        MLFB = imput[0:20]
        Z = "?"
        barvaZ = "#ff0000"
        #print(z)
        #print(MLFB)
        datum = imput[24:26]
        #print(datum)
    
        vc = imput[24:38]
        wc = len(vc)
        #print(wc)


        if wc == 13 :

            cislo = imput[24:28]+"-"+imput[28:32]+"-"+imput[32:34]+"-"+imput[34:37]
            #print(cislo)

        elif wc == 14:     

            cislo = imput[24:29]+"-"+imput[29:33]+"-"+imput[33:35]+"-"+imput[35:38]
            #print(cislo)
    
    elif cd < 37 :
        MLFB = imput[0:18]
        Z = "/"
        barvaZ = "#dddddd"
        #print(MLFB)
        #print(z)
        datum = imput[22:24]
        #print(datum)
        
        vc = imput[22:38]
        wc = len(vc)
        #print(wc)
        if wc == 13 :

            cislo = imput[22:26]+"-"+imput[26:30]+"-"+imput[30:32]+"-"+imput[32:35]
            #print(cislo)

        elif wc == 14:     

            cislo = imput[22:27]+"-"+imput[27:31]+"-"+imput[31:33]+"-"+imput[33:36]
            #print(cislo)
        

    #tvorba indexu 1ft6108a-8
    index = imput[0:6]+imput[15]+imput[7:9]
    print(index)

    #vyhledání indexu v tabulce
    for r in ws.rows:
        if r[0].value == index:
            odmer = r[1].value
            barvaO = "#dddddd"

    try:
        odmer
    except NameError:
        #print("Variable is not defined....!")
        odmer = "?"
        barvaO = "#ff0000"
        print(odmer)
    else:
        #print("Variable is defined.")
        print(odmer)
    
    #vyhledání data v tabulce
    for r in ws.rows:
        if wc == 13 :
            if r[2].value == datum:
                dat_v = r[3].value
                #print(dat_v)

        if wc == 14 :
            if r[4].value == datum:
                dat_v = r[5].value
                #print(dat_v)
    return(MLFB,cislo,odmer,Z,dat_v,barvaO,barvaZ)	

def vyroba2(cislo):
    wb = load_workbook(filename = 'data/data.xlsx')
    ws = wb['nic']
    cislo = cislo.replace("-","")
    cislo = cislo.replace(" ","")
    wc = len(cislo)
    datum = cislo[0:2]

    if not wc == 13|14:
        #print("nic")
        dat_v = "leden 2000"
    
    #vyhledání data v tabulce
    for r in ws.rows:
        if wc == 13 :
            if r[2].value == datum:
                dat_v = r[3].value
                #print(dat_v)

        if wc == 14 :
            if r[4].value == datum:
                dat_v = r[5].value
                #print(dat_v)
    return(dat_v)	
